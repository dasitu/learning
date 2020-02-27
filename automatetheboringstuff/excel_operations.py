import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheets = wb.worksheets

for sheet in sheets:
    print("current sheet name: " + sheet.title)
    for row in range(1, sheet.max_row + 1):
        print(f'====== row:{row} ======')
        for col in range(1, sheet.max_column + 1):
            print(f'||    col: {col}    ||')
            print(sheet.cell(row, col).value)

